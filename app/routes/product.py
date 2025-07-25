# app/routes/product.py
import os
from datetime import datetime, date
from io import BytesIO
from flask import render_template, request, redirect, url_for, flash, send_file, current_app, jsonify
from sqlalchemy import case
from fpdf import FPDF
from openpyxl import Workbook
from dateutil.relativedelta import relativedelta
import json

from . import bp
from ..extensions import db
from ..models import Product, InTransitInvoice, InTransitOrder, Currency

@bp.route('/update-minimum-stock', methods=['POST'])
def update_minimum_stock():
    product_id = request.form.get('product_id')
    product = Product.query.get(product_id)
    if product:
        min_stock_str = request.form.get('minimum_stock')
        product.minimum_stock = int(min_stock_str) if min_stock_str and min_stock_str.isdigit() else None
        db.session.commit()
        # ▼▼▼ Повертаємо JSON-відповідь ▼▼▼
        return jsonify({'status': 'success', 'message': f'Мінімальний залишок для {product.sku} оновлено.'})
    return jsonify({'status': 'error', 'message': 'Товар не знайдено.'}), 404


@bp.route('/update-delivery-time', methods=['POST'])
def update_delivery_time():
    product_id = request.form.get('product_id')
    product = Product.query.get(product_id)
    if product:
        delivery_time_str = request.form.get('delivery_time')
        product.delivery_time = int(delivery_time_str) if delivery_time_str and delivery_time_str.isdigit() else 100
        db.session.commit()
        # ▼▼▼ Повертаємо JSON-відповідь ▼▼▼
        return jsonify({'status': 'success', 'message': f'Термін доставки для {product.sku} оновлено.'})
    return jsonify({'status': 'error', 'message': 'Товар не знайдено.'}), 404

@bp.route('/create-in-transit-invoice', methods=['POST'])
def create_in_transit_invoice():
    try:
        data = request.get_json()
        products_data = data.get('products')
        invoice_number = data.get('invoice_number')
        invoice_date_str = data.get('invoice_date')
        # Поле "Коментар" більше не обробляється
        # comment = data.get('comment')
        
        currency_code = data.get('currency_code')
        if not currency_code:
            return jsonify({'status': 'error', 'message': 'Будь ласка, оберіть валюту для накладної.'}), 400

        if not products_data:
            return jsonify({'status': 'error', 'message': 'Не обрано жодного товару.'}), 400

        invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d').date() if invoice_date_str else date.today()

        new_invoice = InTransitInvoice(
            invoice_number=invoice_number,
            invoice_date=invoice_date,
            # comment=comment, # <-- Видалено
            currency_code=currency_code
        )
        db.session.add(new_invoice)
        db.session.flush()

        for item in products_data:
            product = Product.query.get(item['id'])
            quantity = int(item['quantity'])
            cost_price = float(item['cost_price']) if item.get('cost_price') else product.vendor_price
            
            if product and quantity > 0:
                new_order_item = InTransitOrder(
                    invoice_id=new_invoice.id,
                    product_id=product.id,
                    quantity=quantity,
                    cost_price=cost_price
                )
                db.session.add(new_order_item)
                product.in_transit_quantity += quantity
        
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Накладну успішно створено.'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating in-transit invoice: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'Помилка сервера: {e}'}), 500
# --- КІНЕЦЬ ЗМІН ---

@bp.route('/add-in-transit-form/<int:product_id>')
def add_in_transit_form(product_id):
    product = Product.query.get_or_404(product_id)
    return render_template('add_in_transit_form.html', product=product)

# --- ПОЧАТОК ЗМІН: Нові функції для роботи з накладними ---

@bp.route('/in-transit/invoice/delete/<int:invoice_id>', methods=['POST'])
def delete_in_transit_invoice(invoice_id):
    invoice = InTransitInvoice.query.get_or_404(invoice_id)
    try:
        # Віднімаємо кількість товарів перед видаленням
        for item in invoice.items:
            product = item.product
            if product.in_transit_quantity >= item.quantity:
                product.in_transit_quantity -= item.quantity
            else:
                product.in_transit_quantity = 0

        db.session.delete(invoice)
        db.session.commit()
        flash(f"Накладну №{invoice.id} та всі пов'язані товари видалено.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Помилка під час видалення накладної: {e}", 'danger')
        
    return redirect(url_for('main.in_transit_view'))

# --- КІНЕЦЬ ЗМІН ---
# --- ПОЧАТОК ЗМІН: Нова функція для видалення позиції з накладної ---
@bp.route('/in-transit/item/delete/<int:item_id>', methods=['POST'])
def delete_in_transit_item(item_id):
    item_to_delete = InTransitOrder.query.get_or_404(item_id)
    product = item_to_delete.product
    invoice = item_to_delete.invoice
    invoice_id_redirect = invoice.id

    # Віднімаємо кількість з товару, що "в дорозі"
    if product.in_transit_quantity >= item_to_delete.quantity:
        product.in_transit_quantity -= item_to_delete.quantity
    else:
        product.in_transit_quantity = 0

    try:
        db.session.delete(item_to_delete)
        
        # Перевіряємо, чи залишилися в накладній інші товари
        if not invoice.items.count():
            # Якщо товарів не залишилося, видаляємо і саму накладну
            db.session.delete(invoice)
            flash(f"Останній товар було видалено, тому накладну №{invoice.id} також видалено.", 'warning')
            db.session.commit()
            return redirect(url_for('main.in_transit_view'))
        
        db.session.commit()
        flash(f"Позицію '{product.name}' було видалено з накладної.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Помилка під час видалення позиції: {e}", "danger")
        
    return redirect(url_for('main.in_transit_detail_view', invoice_id=invoice_id_redirect))
# --- КІНЕЦЬ ЗМІН ---

@bp.route('/add-in-transit', methods=['POST'])
def add_in_transit():
    # ... (код цієї функції без змін) ...
    product_id = request.form.get('product_id')
    quantity = request.form.get('quantity')
    product = Product.query.get(product_id)
    if not product or not quantity.isdigit() or int(quantity) <= 0:
        flash("Помилка: Некоректні дані.", "danger")
        return redirect(url_for('main.index'))
    product.in_transit_quantity += int(quantity)
    arrival_date_str = request.form.get('arrival_date')
    arrival_date = datetime.strptime(arrival_date_str, '%Y-%m-%d').date() if arrival_date_str else None
    new_order = InTransitOrder(
        product_id=product.id,
        quantity=int(quantity),
        order_number=request.form.get('order_number'),
        arrival_date=arrival_date
    )
    db.session.add(new_order)
    db.session.commit()
    flash(f"Додано {quantity} од. товару '{product.name}' в дорозі.", "success")
    return redirect(request.referrer or url_for('main.index'))


@bp.route('/bulk-actions', methods=['POST'])
def bulk_actions():
    # ... (код цієї функції без змін) ...
    action = request.form.get('action')
    product_ids_str = request.form.getlist('product_ids')
    display_currency_code = request.form.get('display_currency', 'UAH')
    if not product_ids_str:
        flash('Помилка: товари не обрано.', 'warning')
        return redirect(url_for('main.index'))
    product_ids = [int(pid) for pid in product_ids_str]
    products = Product.query.filter(Product.id.in_(product_ids)).all()
    currency = Currency.query.filter_by(code=display_currency_code).first()
    rate = currency.rate if currency and currency.rate > 0 else 1.0
    product_map = {product.id: product for product in products}
    sorted_products = [product_map[pid] for pid in product_ids if pid in product_map]
    if action == 'export_xls':
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(['Артикул', 'Назва товару', 'Ціна', 'Залишок на складі'])
        for product in sorted_products:
            converted_price = product.price / rate
            formatted_price = f'{converted_price:.2f}'.replace('.', ',')
            ws.append([product.sku, product.name, formatted_price, product.stock])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name='products.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif action == 'export_xls_plus':
        wb = Workbook()
        ws = wb.active
        ws.title = "Export_XLS_Plus"
        ws.append(['ID товару/послуги', 'Назва (UA)', 'SKU', 'Ціна', 'Ціна - Валюта', 'Знижка', 'Ціна зі знижкою', 'Період знижки від', 'Період знижки до'])
        today = date.today()
        three_years_later = today + relativedelta(years=3)
        for product in sorted_products:
            price_uah = product.price
            ws.append([product.vendor_code, product.name, product.sku, f'{price_uah:.2f}', 'UAH', '', '', today.strftime('%d.%m.%Y'), three_years_later.strftime('%d.%m.%Y')])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"export_plus_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif action == 'export_pdf':
        class PDF(FPDF):
            def header(self):
                if self.page_no() == 1:
                    self.set_font('DejaVu', 'B', 14)
                    self.cell(0, 10, 'Список товарів', 0, 1, 'C')
                    self.ln(8)
                self.set_font('DejaVu', 'B', 9)
                self.set_fill_color(240, 240, 240)
                self.cell(self.col_widths["num"], 8, '#', 1, 0, 'C', 1)
                self.cell(self.col_widths["sku"], 8, 'Артикул', 1, 0, 'C', 1)
                self.cell(self.col_widths["name"], 8, 'Назва товару', 1, 0, 'C', 1)
                self.cell(self.col_widths["price"], 8, f'Ціна, {display_currency_code}', 1, 0, 'C', 1)
                self.cell(self.col_widths["stock"], 8, 'Залишок', 1, 0, 'C', 1)
                self.cell(self.col_widths["note"], 8, 'Нотатка', 1, 1, 'C', 1)
            def footer(self):
                self.set_y(-15)
                self.set_font('DejaVu', '', 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 10, f'Сторінка {self.page_no()}', 0, 0, 'L')
        pdf = PDF('P', 'mm', 'A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        page_width = pdf.w - 2 * pdf.l_margin
        pdf.col_widths = {"num": 10, "sku": 40, "price": 25, "stock": 25, "note": 25}
        pdf.col_widths["name"] = page_width - sum(pdf.col_widths.values())
        font_path = os.path.join(current_app.root_path, 'static', 'fonts', 'DejaVuSans.ttf')
        font_path_bold = os.path.join(current_app.root_path, 'static', 'fonts', 'DejaVuSans-Bold.ttf')
        if not os.path.exists(font_path) or not os.path.exists(font_path_bold):
             flash(f'Помилка: Файли шрифтів не знайдено в "app/static/fonts/".', 'danger')
             return redirect(url_for('main.index'))
        pdf.add_font('DejaVu', '', font_path)
        pdf.add_font('DejaVu', 'B', font_path_bold)
        pdf.add_page()
        default_font_size = 8
        pdf.set_font('DejaVu', '', default_font_size)
        name_font_size = 6
        name_cell_height = 3
        for i, product in enumerate(sorted_products):
            pdf.set_font_size(name_font_size)
            row_height = (len(pdf.multi_cell(pdf.col_widths["name"], name_cell_height, product.name, border=0, dry_run=True, output="LINES"))) * name_cell_height
            pdf.set_font_size(default_font_size)
            if pdf.get_y() + row_height > pdf.page_break_trigger:
                pdf.add_page()
            y_before_row = pdf.get_y()
            pdf.cell(pdf.col_widths["num"], row_height, str(i + 1), border=1, align='C')
            pdf.cell(pdf.col_widths["sku"], row_height, product.sku, border=1, align='L')
            x_after_name = pdf.get_x() + pdf.col_widths["name"]
            pdf.set_font_size(name_font_size)
            pdf.multi_cell(pdf.col_widths["name"], name_cell_height, product.name, border=1, align='L')
            pdf.set_xy(x_after_name, y_before_row)
            pdf.set_font_size(default_font_size)
            converted_price = product.price / rate
            pdf.cell(pdf.col_widths["price"], row_height, f"{converted_price:.2f}", border=1, align='C')
            pdf.cell(pdf.col_widths["stock"], row_height, str(product.stock), border=1, align='C')
            pdf.cell(pdf.col_widths["note"], row_height, '', border=1, ln=1)
        buffer = pdf.output()
        return send_file(BytesIO(buffer), as_attachment=True, download_name='products.pdf', mimetype='application/pdf')
    flash('Невідома дія.', 'danger')
    return redirect(url_for('main.index'))


@bp.route('/export/goods-receipt', methods=['POST'])
def export_goods_receipt():
    # ... (код цієї функції без змін) ...
    product_ids_str = request.form.get('product_ids')
    export_currency_code = request.form.get('currency')
    if not product_ids_str or not export_currency_code:
        flash('Помилка: не обрано товари або валюту.', 'danger')
        return redirect(url_for('main.index'))
    product_ids = [int(pid) for pid in product_ids_str.split(',')]
    ordering = case({pid: index for index, pid in enumerate(product_ids)}, value=Product.id)
    products = Product.query.filter(Product.id.in_(product_ids)).order_by(ordering).all()
    export_currency = Currency.query.filter_by(code=export_currency_code).first()
    if not export_currency:
        flash(f'Помилка: валюту {export_currency_code} не знайдено.', 'danger')
        return redirect(url_for('main.index'))
    wb = Workbook()
    ws = wb.active
    ws.title = "Надходження"
    ws.append(['ID', 'SKU', 'Назва', 'К-ть', 'Ціна закупки'])
    for product in products:
        price = product.vendor_price if product.vendor_price else 0.0
        if export_currency.code != 'UAH':
            if export_currency.rate > 0:
                price = price / export_currency.rate
            else:
                price = 0.0
        formatted_price = f'{price:.4f}'.replace('.', ',')
        ws.append([product.vendor_code, product.sku, product.name, '', formatted_price])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"goods_receipt_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# --- ПОЧАТОК ЗМІН: Нова функція для видалення ---
@bp.route('/delete-in-transit/<int:order_id>', methods=['POST'])
def delete_in_transit_order(order_id):
    order_to_delete = InTransitOrder.query.get_or_404(order_id)
    product = order_to_delete.product
    
    # Віднімаємо кількість з товару, що "в дорозі"
    if product.in_transit_quantity >= order_to_delete.quantity:
        product.in_transit_quantity -= order_to_delete.quantity
    else:
        product.in_transit_quantity = 0

    try:
        db.session.delete(order_to_delete)
        db.session.commit()
        flash(f"Замовлення для товару '{product.name}' успішно видалено.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Помилка під час видалення замовлення: {e}", "danger")
        
    return redirect(request.referrer or url_for('main.in_transit_view'))
# --- КІНЕЦЬ ЗМІН ---