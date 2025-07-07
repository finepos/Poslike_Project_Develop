import requests
import json
from datetime import datetime
from flask import render_template, current_app, flash, redirect, url_for, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
from datetime import datetime, date
from dateutil.relativedelta import relativedelta # Додайте цей імпорт

from . import bp
from ..models import Setting, Printer, PrintJob, Currency
from ..extensions import db
from ..utils import natural_sort_key
from ..printing import generate_zpl_code

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException

def _get_authenticated_session(force_login=False):
    # ... (код цієї функції залишається без змін) ...
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
    })

    cookies_json_setting = Setting.query.get('salesdrive_cookies_json')

    if not force_login and cookies_json_setting and cookies_json_setting.value:
        try:
            saved_cookies = json.loads(cookies_json_setting.value)
            for cookie in saved_cookies:
                session.cookies.set(cookie['name'], cookie['value'])
            return session, None
        except (json.JSONDecodeError, TypeError):
            pass

    domain_setting = Setting.query.get('salesdrive_domain')
    login_setting = Setting.query.get('salesdrive_login')
    password_setting = Setting.query.get('salesdrive_password')

    if not all([domain_setting, login_setting, password_setting]) or not all([domain_setting.value, login_setting.value, password_setting.value]):
        return None, 'Дані для входу (домен, логін, пароль) не налаштовані.'

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    driver = None
    try:
        driver = webdriver.Chrome(options=chrome_options)
        login_url = "https://auth.salesdrive.me/auth/login/?lang=ua"
        driver.get(login_url)
        wait = WebDriverWait(driver, 15)

        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="subdomain"]'))).send_keys(domain_setting.value)
        driver.find_element(By.XPATH, '//*[@id="login"]').send_keys(login_setting.value)
        driver.find_element(By.XPATH, '//*[@id="form-login"]/div[3]/div/input').send_keys(password_setting.value)
        driver.find_element(By.XPATH, '//*[@id="btn-login"]').click()

        try:
            wait.until(EC.url_contains(domain_setting.value))
        except TimeoutException:
            return None, 'Не вдалося увійти. Перевірте домен, логін та пароль (Таймаут).'

        if 'auth.salesdrive.me' in driver.current_url:
            return None, 'Не вдалося увійти. Перевірте домен, логін та пароль.'

        selenium_cookies = driver.get_cookies()

        cookies_json = json.dumps(selenium_cookies)
        if cookies_json_setting:
            cookies_json_setting.value = cookies_json
        else:
            db.session.add(Setting(key='salesdrive_cookies_json', value=cookies_json))
        db.session.commit()

        for cookie in selenium_cookies:
            session.cookies.set(cookie['name'], cookie['value'])

        return session, None

    except WebDriverException as e:
        current_app.logger.error(f"Помилка Selenium WebDriver: {e}")
        return None, "Помилка запуску браузера. Перевірте налаштування Selenium WebDriver."
    except Exception as e:
        current_app.logger.error(f"Неочікувана помилка під час входу: {e}")
        return None, f"Неочікувана помилка під час входу: {e}"
    finally:
        if driver:
            driver.quit()


@bp.route('/salesdrive', methods=['GET'])
def salesdrive_index():
    # ... (код цієї функції залишається без змін) ...
    page = request.args.get('page', 1, type=int)
    printers = Printer.query.order_by(Printer.is_default.desc(), Printer.name).all()
    documents = []
    pagination_info = None

    try:
        domain_setting = Setting.query.get('salesdrive_domain')
        if not domain_setting or not domain_setting.value:
            flash('Інтеграція з SalesDrive не налаштована.', 'warning')
            return render_template('salesdrive_list.html', documents=None, pagination_info=None, printers=printers)

        session, error = _get_authenticated_session()
        if error:
            raise Exception(error)

        base_docs_url = f"https://{domain_setting.value}.salesdrive.me/document-arrival-product/index/?active=1&expand=documentItems,documentItems.product,documentItems.unit&formId=1&page={page}"
        response = session.get(base_docs_url, timeout=15)

        if 'auth.salesdrive.me' in response.url:
            flash('Cookie застаріли. Виконується повторний вхід...', 'warning')
            session, error = _get_authenticated_session(force_login=True)
            if error:
                raise Exception(error)

            response = session.get(base_docs_url, timeout=15)
            if 'auth.salesdrive.me' in response.url:
                 raise Exception("Повторна спроба входу також не вдалася.")

        response.raise_for_status()
        data = response.json()
        documents = data.get('data', [])

        if not documents and page > 1:
            flash('Документи на цій сторінці не знайдено. Можливо, це кінець списку.', 'info')

        pagination_info = {
            'current_page': page,
            'has_next': len(documents) > 0,
            'has_prev': page > 1
        }

    except Exception as e:
        flash(f'Сталася помилка: {e}', 'danger')

    return render_template('salesdrive_list.html', documents=documents, pagination_info=pagination_info, printers=printers)


# ▼▼▼ ПОЧАТОК: ПОВНІСТЮ ЗАМІНІТЬ ІСНУЮЧУ ФУНКЦІЮ ▼▼▼
@bp.route('/salesdrive/document/<int:doc_id>')
def salesdrive_document_detail(doc_id):
    """
    Відображає детальну інформацію про прихідну накладну,
    включаючи назву валюти закупки.
    """
    try:
        printers = Printer.query.order_by(Printer.is_default.desc(), Printer.name).all()
        printers_json = json.dumps([p.to_dict() for p in printers])

        session, error = _get_authenticated_session()
        if error:
            raise Exception(error)

        domain = Setting.query.get('salesdrive_domain').value
        detail_url = f"https://{domain}.salesdrive.me/document-arrival-product/{doc_id}/?activeAccount=1&formId=1&id={doc_id}"
        headers = {'Accept': 'application/json, text/plain, */*'}

        response = session.get(detail_url, headers=headers, timeout=15)
        if 'auth.salesdrive.me' in response.url: # Перевірка на застарілі куки
            session, error = _get_authenticated_session(force_login=True)
            if error: raise Exception(error)
            response = session.get(detail_url, headers=headers, timeout=15)
        response.raise_for_status()

        detail_data = response.json()
        document = detail_data.get('response', {}).get('item')
        if not document:
            raise Exception("Структура відповіді API не містить даних про документ.")

        # Визначаємо валюту закупки з даних API
        cost_currency_code = 'N/A'
        currency_id = document.get('currencyId')
        currency_options = detail_data.get('response', {}).get('meta', {}).get('fields', {}).get('currencyId', {}).get('options', [])
        
        # Створюємо словник для пошуку { value: text }
        currency_map = {str(opt['value']): opt['text'] for opt in currency_options}
        
        # Знаходимо текстове представлення валюти
        cost_currency_code = currency_map.get(str(currency_id), 'N/A')

        return render_template('salesdrive_document_detail.html',
                               document=document,
                               cost_currency_code=cost_currency_code,
                               printers=printers,
                               printers_json=printers_json,
                               doc_id=doc_id)

    except Exception as e:
        current_app.logger.error(f"Помилка завантаження деталей документа: {e}", exc_info=True)
        flash(f'Сталася помилка при завантаженні деталей документа: {e}', 'danger')
        return redirect(url_for('main.salesdrive_index'))
# ▲▲▲ КІНЕЦЬ ЗАМІНИ ФУНКЦІЇ ▲▲▲


@bp.route('/salesdrive/print-invoice', methods=['POST'])
def salesdrive_print_invoice():
    # ... (код цієї функції залишається без змін) ...
    try:
        doc_id = request.form.get('doc_id')
        printer_id = request.form.get('printer_id')
        print_option = request.form.get('print_option')

        if not all([doc_id, printer_id, print_option]):
            return jsonify({'status': 'error', 'message': 'Не вказано всі необхідні параметри для друку.'}), 400

        session, error = _get_authenticated_session()
        if error:
            return jsonify({'status': 'error', 'message': f'Помилка автентифікації: {error}'}), 500

        domain = Setting.query.get('salesdrive_domain').value
        detail_url = f"https://{domain}.salesdrive.me/document-arrival-product/{doc_id}/?activeAccount=1&formId=1&id={doc_id}"
        headers = {'Accept': 'application/json'}
        response = session.get(detail_url, headers=headers, timeout=15)
        response.raise_for_status()

        document = response.json().get('response', {}).get('item')
        if not document or not document.get('documentItems'):
            return jsonify({'status': 'error', 'message': 'Накладна не містить товарів.'}), 404

        printer = Printer.query.get(printer_id)
        if not printer:
            return jsonify({'status': 'error', 'message': 'Принтер не знайдено.'}), 404

        items_to_print = sorted(document['documentItems'], key=lambda x: natural_sort_key(x.get('description')))

        full_zpl_code = ""
        total_labels = 0

        for item in items_to_print:
            item_product_data = item.get('product', {})
            product_name_for_print = item_product_data.get('nameTranslate', item.get('description', ''))

            try:
                price_from_invoice = float(item.get('price', 0))
            except (ValueError, TypeError):
                price_from_invoice = 0.0

            xml_data_for_print = {
                'product_sku': item_product_data.get('sku', ''),
                'product_name': product_name_for_print,
                'product_price': f"{price_from_invoice:.4f}",
                'product_url': item_product_data.get('href', ''),
            }

            pseudo_product = type('obj', (object,), {
                'id': item.get('productId'),
                'sku': item_product_data.get('sku', ''),
                'name': product_name_for_print,
                'price': price_from_invoice,
                'xml_data': json.dumps(xml_data_for_print, ensure_ascii=False)
            })

            quantity = 0
            if print_option == 'by_quantity':
                quantity = int(float(item.get('count', 0)))
            elif print_option == 'one_per_item':
                quantity = 1

            if quantity > 0:
                full_zpl_code += generate_zpl_code(printer, pseudo_product, quantity=quantity)
                total_labels += quantity

        if not full_zpl_code:
            return jsonify({'status': 'error', 'message': 'Не вдалося створити етикетки для друку.'}), 500

        new_job = PrintJob(printer_id=printer.id, zpl_code=full_zpl_code)
        db.session.add(new_job)
        db.session.commit()

        message = f"Створено завдання на друк ({total_labels} етикеток) для принтера '{printer.name}'."
        return jsonify({'status': 'success', 'message': message})

    except Exception as e:
        current_app.logger.error(f"Помилка під час створення завдання на друк: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'Помилка сервера: {e}'}), 500


# ▼▼▼ ПОЧАТОК: ПОВНІСТЮ ЗАМІНІТЬ ІСНУЮЧУ ФУНКЦІЮ НА ЦЮ ▼▼▼
@bp.route('/salesdrive/export-xls', methods=['POST'])
def salesdrive_export_xls():
    """
    Генерує XLS файл, коректно обробляючи відсоткові та фіксовані знижки,
    та встановлює відповідний формат комірок.
    """
    selected_products_json = request.form.getlist('selected_products')
    doc_id = request.form.get('doc_id')

    if not selected_products_json or not doc_id:
        flash('Не обрано товари для експорту або відсутній ID документа.', 'warning')
        return redirect(request.referrer)

    try:
        # --- Отримання даних з SalesDrive ---
        session, error = _get_authenticated_session()
        if error: raise Exception(error)
        
        domain = Setting.query.get('salesdrive_domain').value
        detail_url = f"https://{domain}.salesdrive.me/document-arrival-product/{doc_id}/?activeAccount=1&formId=1&id={doc_id}"
        response = session.get(detail_url, headers={'Accept': 'application/json'}, timeout=15)
        response.raise_for_status()
        
        detail_data = response.json()
        currency_options = detail_data.get('response', {}).get('meta', {}).get('fields', {}).get('currencyId', {}).get('options', [])
        currency_name_map = {str(opt['value']): opt['text'] for opt in currency_options}
        # --- Кінець отримання даних ---

        wb = Workbook()
        ws = wb.active
        ws.title = "SalesDrive Export"
        ws.append([
            'ID товару/послуги', 'Назва (UA)', 'SKU', 'Ціна', 'Ціна - Валюта',
            'Знижка', 'Ціна зі знижкою', 'Період знижки від', 'Період знижки до',
            'Собівартість', 'Собівартість - Валюта', 'К-ть'
        ])

        today = date.today()
        three_years_later = today + relativedelta(years=3)

        for product_json_str in selected_products_json:
            item_data = json.loads(product_json_str)
            product_attrs = item_data.get('product', {})

            price = float(product_attrs.get('defaultPrice', 0))
            
            # --- ВИПРАВЛЕНО: Нова логіка розрахунку знижки ---
            is_percent_discount = product_attrs.get('percentDiscount') == '1'
            discount_raw_value = float(product_attrs.get('discount', 0))
            
            price_with_discount = price
            discount_for_cell = 0

            if is_percent_discount and discount_raw_value > 0:
                # Знижка у відсотках
                price_with_discount = price * (1 - discount_raw_value / 100)
                discount_for_cell = discount_raw_value / 100  # Зберігаємо як десятковий дріб (напр. 0.04)
            elif discount_raw_value > 0:
                # Фіксована знижка
                price_with_discount = price - discount_raw_value
                discount_for_cell = discount_raw_value
            # --- Кінець нової логіки ---

            # --- Форматування дат ---
            date_from_raw = product_attrs.get('discountPeriodFrom')
            date_to_raw = product_attrs.get('discountPeriodTo')

            if date_from_raw and len(date_from_raw) >= 10:
                date_from = datetime.strptime(date_from_raw[:10], '%Y-%m-%d').strftime('%d.%m.%Y')
            else:
                date_from = today.strftime('%d.%m.%Y')

            if date_to_raw and len(date_to_raw) >= 10:
                date_to = datetime.strptime(date_to_raw[:10], '%Y-%m-%d').strftime('%d.%m.%Y')
            else:
                date_to = three_years_later.strftime('%d.%m.%Y')
            
            # --- Визначення собівартості та її валюти ---
            cost_price_value = product_attrs.get('costPriceCurrency')
            if not cost_price_value or float(cost_price_value) == 0:
                cost_price_value = product_attrs.get('costPrice', 0)

            cost_price_currency_id_raw = product_attrs.get('costPriceCurrencyId')
            if cost_price_currency_id_raw == 0 or cost_price_currency_id_raw is None:
                cost_price_currency_name = 'UAH'
            else:
                cost_price_currency_name = currency_name_map.get(str(cost_price_currency_id_raw), 'N/A')

            row = [
                product_attrs.get('parameter'),
                product_attrs.get('nameTranslate'),
                product_attrs.get('sku'),
                f"{price:.2f}".replace('.', ','),
                product_attrs.get('defaultPriceCurrency', 'UAH'),
                discount_for_cell, # Вставляємо підготовлене значення
                f"{price_with_discount:.2f}".replace('.', ','),
                date_from,
                date_to,
                f"{float(cost_price_value):.4f}".replace('.', ','),
                cost_price_currency_name,
                item_data.get('count')
            ]
            ws.append(row)
            
            # ВИПРАВЛЕНО: Встановлюємо формат комірки, якщо знижка відсоткова
            if is_percent_discount:
                # Отримуємо комірку 'Знижка' (6-та колонка) щойно доданого рядка
                discount_cell = ws.cell(row=ws.max_row, column=6) 
                discount_cell.number_format = '0.00%'

    except Exception as e:
        current_app.logger.error(f"Помилка під час генерації XLS: {e}", exc_info=True)
        flash(f'Сталася критична помилка під час створення файлу: {e}', 'danger')
        return redirect(request.referrer)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"salesdrive_export_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# ▲▲▲ КІНЕЦЬ ЗАМІНИ ФУНКЦІЇ ▲▲▲